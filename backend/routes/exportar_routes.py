"""
Rutas para exportar datos a Excel
"""
from flask import Blueprint, send_file, jsonify
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from app import db
from models.pedido import Pedido
from models.producto import Producto
from models.cliente import Cliente

bp = Blueprint('exportar', __name__)


def crear_excel_pedidos():
    """Genera archivo Excel con todos los pedidos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    
    # Encabezados
    headers = [
        'ID', 'Fecha Pedido', 'Fecha Entrega', 'Canal', 'N° Shopify',
        'Cliente', 'Teléfono', 'Arreglo', 'Detalles', 'Precio Ramo',
        'Precio Envío', 'Destinatario', 'Mensaje', 'Firma', 'Dirección',
        'Comuna', 'Motivo', 'Estado', 'Pagado'
    ]
    
    # Estilo de encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    pedidos = Pedido.query.order_by(Pedido.fecha_pedido.desc()).all()
    
    for row_num, pedido in enumerate(pedidos, 2):
        ws.cell(row=row_num, column=1, value=pedido.id)
        ws.cell(row=row_num, column=2, value=pedido.fecha_pedido.strftime('%Y-%m-%d') if pedido.fecha_pedido else '')
        ws.cell(row=row_num, column=3, value=pedido.fecha_entrega.strftime('%Y-%m-%d') if pedido.fecha_entrega else '')
        ws.cell(row=row_num, column=4, value=pedido.canal or '')
        ws.cell(row=row_num, column=5, value=pedido.shopify_order_number or '')
        ws.cell(row=row_num, column=6, value=pedido.cliente_nombre or '')
        ws.cell(row=row_num, column=7, value=pedido.cliente_telefono or '')
        ws.cell(row=row_num, column=8, value=pedido.arreglo_pedido or '')
        ws.cell(row=row_num, column=9, value=pedido.detalles_adicionales or '')
        ws.cell(row=row_num, column=10, value=pedido.precio_ramo or 0)
        ws.cell(row=row_num, column=11, value=pedido.precio_envio or 0)
        ws.cell(row=row_num, column=12, value=pedido.destinatario or '')
        ws.cell(row=row_num, column=13, value=pedido.mensaje or '')
        ws.cell(row=row_num, column=14, value=pedido.firma or '')
        ws.cell(row=row_num, column=15, value=pedido.direccion_entrega or '')
        ws.cell(row=row_num, column=16, value=pedido.comuna or '')
        ws.cell(row=row_num, column=17, value=pedido.motivo or '')
        ws.cell(row=row_num, column=18, value=pedido.estado or '')
        ws.cell(row=row_num, column=19, value=pedido.estado_pago or '')
    
    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    return wb


def crear_excel_productos():
    """Genera archivo Excel con todos los productos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Encabezados
    headers = [
        'ID', 'Nombre', 'Descripción', 'Precio Venta', 'Costo Estimado',
        'Margen (%)', 'Vista', 'Tamaño', 'Cuidados', 'Activo'
    ]
    
    # Estilo de encabezados
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    productos = Producto.query.all()
    
    for row_num, producto in enumerate(productos, 2):
        ws.cell(row=row_num, column=1, value=producto.id)
        ws.cell(row=row_num, column=2, value=producto.nombre)
        ws.cell(row=row_num, column=3, value=producto.descripcion or '')
        ws.cell(row=row_num, column=4, value=float(producto.precio_venta) if producto.precio_venta else 0)
        ws.cell(row=row_num, column=5, value=float(producto.costo_estimado) if producto.costo_estimado else 0)
        
        # Calcular margen
        if producto.precio_venta and producto.costo_estimado and producto.precio_venta > 0:
            margen = ((float(producto.precio_venta) - float(producto.costo_estimado)) / float(producto.precio_venta)) * 100
            ws.cell(row=row_num, column=6, value=round(margen, 2))
        else:
            ws.cell(row=row_num, column=6, value=0)
        
        ws.cell(row=row_num, column=7, value=producto.vista or '')
        ws.cell(row=row_num, column=8, value=producto.tamano or '')
        ws.cell(row=row_num, column=9, value=producto.cuidados or '')
        ws.cell(row=row_num, column=10, value='Sí' if producto.activo else 'No')
    
    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    return wb


def crear_excel_clientes():
    """Genera archivo Excel con todos los clientes"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Encabezados
    headers = [
        'ID', 'Nombre', 'Teléfono', 'Email', 'Tipo Cliente',
        'Plazo Pago (días)', 'Total Pedidos', 'Total Gastado', 'Activo'
    ]
    
    # Estilo de encabezados
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    clientes = Cliente.query.all()
    
    for row_num, cliente in enumerate(clientes, 2):
        ws.cell(row=row_num, column=1, value=cliente.id)
        ws.cell(row=row_num, column=2, value=cliente.nombre)
        ws.cell(row=row_num, column=3, value=cliente.telefono or '')
        ws.cell(row=row_num, column=4, value=cliente.email or '')
        ws.cell(row=row_num, column=5, value=cliente.tipo_cliente or '')
        ws.cell(row=row_num, column=6, value=cliente.plazo_pago_dias or 0)
        ws.cell(row=row_num, column=7, value=cliente.total_pedidos or 0)
        ws.cell(row=row_num, column=8, value=cliente.total_gastado or 0)
        ws.cell(row=row_num, column=9, value='Sí' if cliente.activo else 'No')
    
    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    return wb


@bp.route('/pedidos', methods=['GET'])
def exportar_pedidos():
    """Exporta todos los pedidos a Excel"""
    try:
        wb = crear_excel_pedidos()
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con fecha
        filename = f'Pedidos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/productos', methods=['GET'])
def exportar_productos():
    """Exporta todos los productos a Excel"""
    try:
        wb = crear_excel_productos()
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con fecha
        filename = f'Productos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/clientes', methods=['GET'])
def exportar_clientes():
    """Exporta todos los clientes a Excel"""
    try:
        wb = crear_excel_clientes()
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con fecha
        filename = f'Clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

