#!/usr/bin/env python3
"""
Genera Excel demo con estructura de eventos (VERSIÓN 2 CON FALTANTES)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

def crear_excel_eventos():
    """Crea Excel con datos demo de eventos"""
    
    wb = openpyxl.Workbook()
    
    # ========================================
    # HOJA 1: EVENTOS
    # ========================================
    ws_eventos = wb.active
    ws_eventos.title = "01_Eventos"
    
    headers_eventos = [
        'ID Evento', 'Cliente Nombre', 'Cliente Teléfono', 'Cliente Email',
        'Nombre Evento', 'Tipo Evento', 'Fecha Evento', 'Hora Evento',
        'Lugar Evento', 'Cantidad Personas', 'Estado', 'Costo Insumos',
        'Costo Mano Obra', 'Costo Transporte', 'Costo Otros', 'Costo Total',
        'Margen %', 'Precio Propuesta', 'Precio Final', 'Anticipo', 'Saldo',
        'Pagado', 'Insumos Reservados', 'Insumos Descontados', 'Insumos Faltantes',
        'Notas Cotización', 'Fecha Cotización'
    ]
    
    ws_eventos.append(headers_eventos)
    
    # Estilo headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws_eventos[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos de eventos
    eventos_data = [
        # EV001: Boda - En cotización
        ['EV001', 'María González', '+56912345678', 'maria.gonzalez@email.com', 
         'Boda María & Juan', 'Boda', '2025-11-15', '18:00', 
         'Hotel Plaza, Santiago', 150, 'Cotización', 0, 80000, 50000, 0, 130000,
         30, 169000, 0, 0, 169000, False, False, False, False, 
         'Cliente solicita flores blancas y rosadas', '2025-10-23'],
        
        # EV002: Corporativo - Propuesta enviada
        ['EV002', 'Empresa TechCorp', '+56987654321', 'eventos@techcorp.cl', 
         'Aniversario 10 años TechCorp', 'Corporativo', '2025-12-01', '19:30', 
         'Centro de Eventos Espacio Riesco', 200, 'Propuesta Enviada', 380000, 120000, 80000, 50000,
         630000, 25, 787500, 0, 0, 787500, False, False, False, False, 
         'Evento corporativo formal, colores azul y blanco', '2025-10-20'],
        
        # EV003: Cumpleaños - Confirmado
        ['EV003', 'Carolina Pérez', '+56923456789', 'caro.perez@gmail.com', 
         'Cumpleaños 30 de Carolina', 'Cumpleaños', '2025-10-30', '20:00', 
         'Restaurant El Jardín, Providencia', 80, 'Confirmado', 150000, 50000, 30000, 20000,
         250000, 35, 337500, 337500, 100000, 237500, False, True, False, False, 
         'Temática tropical, muchas flores coloridas', '2025-10-15'],
        
        # EV004: Baby Shower - RETIRADO CON FALTANTES ⚠️
        ['EV004', 'Andrea Morales', '+56934567890', 'andrea.morales@hotmail.com', 
         'Baby Shower Sofía', 'Baby Shower', '2025-10-20', '16:00', 
         'Salón de Eventos La Rosaleda', 60, 'Retirado', 95000, 35000, 25000, 10000,
         165000, 30, 214500, 214500, 214500, 0, True, True, True, True, 
         '⚠️ FALTAN: 2 maceteros PE002, 3 velas PE006 - Cliente debe devolver', '2025-10-05'],
        
        # EV005: Graduación - En preparación
        ['EV005', 'Universidad Central', '+56945678901', 'ceremonias@ucentral.cl', 
         'Ceremonia Graduación 2025', 'Graduación', '2025-11-05', '11:00', 
         'Auditorio Universidad Central', 300, 'En Preparación', 280000, 90000, 60000, 40000,
         470000, 20, 564000, 564000, 200000, 364000, False, True, True, False, 
         'Arreglos elegantes para escenario y pasillo', '2025-10-10']
    ]
    
    for evento in eventos_data:
        ws_eventos.append(evento)
    
    # ========================================
    # HOJA 2: INSUMOS DE EVENTOS
    # ========================================
    ws_insumos = wb.create_sheet("02_Insumos_Eventos")
    
    headers_insumos = [
        'Evento ID', 'Tipo Insumo', 'ID/Código Insumo', 'Nombre Insumo',
        'Cantidad', 'Costo Unitario', 'Costo Total', 'Reservado', 
        'Descontado Stock', 'Devuelto', 'Cantidad Faltante', 'Notas'
    ]
    
    ws_insumos.append(headers_insumos)
    
    for cell in ws_insumos[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Insumos de eventos
    insumos_data = [
        # EV003: Cumpleaños (confirmado, reservado)
        ['EV003', 'producto_evento', 'PE001', 'Mantel Redondo Blanco 3m', 10, 5000, 50000, True, False, False, 0, ''],
        ['EV003', 'producto_evento', 'PE005', 'Vela Pilar Grande Blanca', 20, 2000, 40000, True, False, False, 0, ''],
        ['EV003', 'flor', 'FL001', 'Rosa Roja Premium', 100, 800, 80000, True, False, False, 0, 'Para centros de mesa'],
        ['EV003', 'contenedor', 'CT001', 'Florero Vidrio Redondo Pequeño', 10, 3000, 30000, True, False, False, 0, ''],
        ['EV003', 'otro', '', 'Mano de Obra - Decoración', 1, 50000, 50000, False, False, False, 0, '6 horas de trabajo'],
        ['EV003', 'otro', '', 'Transporte y Montaje', 1, 30000, 30000, False, False, False, 0, 'Camioneta + ayudante'],
        
        # EV004: Baby Shower (FINALIZADO CON FALTANTES ⚠️)
        ['EV004', 'producto_evento', 'PE002', 'Macetero Terracota Mediano', 5, 4000, 20000, True, True, False, 2, '⚠️ FALTAN 2 maceteros'],
        ['EV004', 'producto_evento', 'PE006', 'Vela Pilar Pequeña Rosada', 10, 1500, 15000, True, True, False, 3, '⚠️ FALTAN 3 velas'],
        ['EV004', 'producto_evento', 'PE001', 'Mantel Redondo Blanco 3m', 6, 5000, 30000, True, True, True, 0, 'Devueltos OK'],
        ['EV004', 'flor', 'FL005', 'Clavel Blanco', 80, 400, 32000, True, True, True, 0, 'Usados en arreglos'],
        ['EV004', 'contenedor', 'CT003', 'Canasto Mimbre Natural', 8, 2500, 20000, True, True, True, 0, 'Devueltos OK'],
        ['EV004', 'otro', '', 'Mano de Obra - Montaje', 1, 35000, 35000, False, True, True, 0, '4 horas'],
        ['EV004', 'otro', '', 'Transporte', 1, 25000, 25000, False, True, True, 0, 'Ida y retiro'],
        
        # EV005: Graduación (en preparación)
        ['EV005', 'producto_evento', 'PE007', 'Arco Floral Grande', 2, 50000, 100000, True, True, False, 0, 'Para escenario'],
        ['EV005', 'producto_evento', 'PE001', 'Mantel Redondo Blanco 3m', 30, 5000, 150000, True, True, False, 0, ''],
        ['EV005', 'flor', 'FL001', 'Rosa Roja Premium', 200, 800, 160000, True, True, False, 0, 'Para arreglos principales'],
        ['EV005', 'flor', 'FL002', 'Lirio Blanco', 100, 1200, 120000, True, True, False, 0, 'Complemento arreglos'],
        ['EV005', 'contenedor', 'CT001', 'Florero Vidrio Redondo Pequeño', 40, 3000, 120000, True, True, False, 0, ''],
        ['EV005', 'otro', '', 'Mano de Obra - Equipo Completo', 1, 90000, 90000, False, True, False, 0, '2 personas x 5 horas'],
        ['EV005', 'otro', '', 'Transporte Camión Grande', 1, 60000, 60000, False, True, False, 0, 'Incluye ayudantes'],
    ]
    
    for insumo in insumos_data:
        ws_insumos.append(insumo)
    
    # ========================================
    # HOJA 3: PRODUCTOS DE EVENTOS
    # ========================================
    ws_productos = wb.create_sheet("03_Productos_Evento")
    
    headers_productos = [
        'Código', 'Nombre', 'Categoría', 'Stock Total', 'En Evento',
        'Disponible', 'Costo Compra', 'Costo Alquiler', 'Descripción',
        'Medidas', 'Color', 'Material'
    ]
    
    ws_productos.append(headers_productos)
    
    for cell in ws_productos[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    productos_data = [
        # Manteles
        ['PE001', 'Mantel Redondo Blanco 3m', 'Mantelería', 50, 10, 40, 8000, 5000, 'Mantel redondo blanco de tela', '3m diámetro', 'Blanco', 'Tela'],
        ['PE002', 'Macetero Terracota Mediano', 'Contenedores', 20, 5, 13, 6000, 4000, 'Macetero de terracota artesanal', '25cm alto', 'Natural', 'Terracota'],
        ['PE003', 'Mantel Rectangular Marfil', 'Mantelería', 40, 0, 40, 10000, 6000, 'Mantel rectangular marfil', '2m x 3m', 'Marfil', 'Tela'],
        ['PE004', 'Mantel Cuadrado Negro', 'Mantelería', 30, 0, 30, 9000, 5500, 'Mantel cuadrado elegante', '2m x 2m', 'Negro', 'Tela'],
        
        # Velas
        ['PE005', 'Vela Pilar Grande Blanca', 'Iluminación', 100, 20, 80, 3000, 2000, 'Vela pilar decorativa grande', '15cm alto', 'Blanco', 'Cera'],
        ['PE006', 'Vela Pilar Pequeña Rosada', 'Iluminación', 80, 10, 67, 2000, 1500, 'Vela pilar pequeña rosada', '10cm alto', 'Rosado', 'Cera'],
        
        # Decoración
        ['PE007', 'Arco Floral Grande', 'Estructura', 5, 2, 3, 80000, 50000, 'Arco metálico para flores', '2.5m x 2m', 'Blanco', 'Metal'],
        ['PE008', 'Triángulo Decorativo Largo', 'Estructura', 10, 0, 10, 45000, 30000, 'Triángulo decorativo metálico', '1.8m alto', 'Dorado', 'Metal'],
        ['PE009', 'Base para Centros de Mesa', 'Soporte', 60, 0, 60, 5000, 3000, 'Base espejo circular', '30cm diámetro', 'Espejo', 'Vidrio'],
        ['PE010', 'Cortina de Luces LED', 'Iluminación', 15, 0, 15, 25000, 15000, 'Cortina LED blanco cálido', '3m x 3m', 'Blanco', 'Cable LED'],
        ['PE011', 'Letras Decorativas Grandes', 'Decoración', 8, 0, 8, 35000, 20000, 'Letras iluminadas LOVE/HAPPY', '80cm alto', 'Blanco', 'MDF'],
        ['PE012', 'Mesa Auxiliar Plegable', 'Mobiliario', 20, 0, 20, 15000, 8000, 'Mesa plegable para buffet', '1.5m x 0.8m', 'Blanco', 'Metal/Plástico'],
    ]
    
    for producto in productos_data:
        # Calcular disponible
        stock = producto[3]
        en_evento = producto[4]
        disponible = stock - en_evento
        producto[5] = disponible
        ws_productos.append(producto)
    
    # ========================================
    # HOJA 4: INSTRUCCIONES
    # ========================================
    ws_instrucciones = wb.create_sheet("INSTRUCCIONES")
    
    instrucciones = [
        ['SISTEMA DE GESTIÓN DE EVENTOS - LAS LIRA'],
        [''],
        ['📋 ESTRUCTURA DEL ARCHIVO'],
        [''],
        ['Hoja 1: 01_Eventos'],
        ['  • Contiene todos los eventos con información completa'],
        ['  • Estados: Cotización → Propuesta → Confirmado → En Preparación → En Evento → Finalizado → Retirado'],
        ['  • Cada evento tiene ID único (EV001, EV002, etc.)'],
        [''],
        ['Hoja 2: 02_Insumos_Eventos'],
        ['  • Detalle de todos los insumos utilizados por evento'],
        ['  • Tipos de insumo: flor, contenedor, producto, producto_evento, otro'],
        ['  • Control de reserva, descuento de stock y devolución'],
        ['  • ⚠️ Cantidad Faltante: Si >0 significa que NO fue devuelto'],
        [''],
        ['Hoja 3: 03_Productos_Evento'],
        ['  • Catálogo de productos específicos para eventos'],
        ['  • Incluye: manteles, velas, arcos, estructuras, iluminación'],
        ['  • Control de stock: Total, En Evento, Disponible'],
        ['  • Costo Compra vs Costo Alquiler'],
        [''],
        ['🔴 EVENTO CON INSUMOS FALTANTES - EJEMPLO'],
        [''],
        ['EV004 - Baby Shower Sofía (Andrea Morales)'],
        ['  • Estado: Retirado'],
        ['  • Insumos Faltantes: TRUE ⚠️'],
        ['  • Faltantes específicos:'],
        ['    - 2 Maceteros Terracota Mediano (PE002)'],
        ['    - 3 Velas Pilar Pequeña Rosada (PE006)'],
        ['  • Acción: Cliente debe devolver los insumos faltantes'],
        ['  • Sistema muestra advertencia en la interfaz'],
        [''],
        ['📊 FLUJO DE TRABAJO'],
        [''],
        ['1. COTIZACIÓN'],
        ['   • Crear evento con datos básicos del cliente'],
        ['   • Agregar insumos estimados'],
        ['   • Calcular costo total + margen → precio propuesta'],
        [''],
        ['2. PROPUESTA ENVIADA'],
        ['   • Cambiar estado cuando se envía presupuesto al cliente'],
        [''],
        ['3. CONFIRMADO'],
        ['   • Cliente acepta la propuesta'],
        ['   • ACCIÓN: Reservar insumos (cantidad_en_evento)'],
        ['   • Insumos Reservados = TRUE'],
        [''],
        ['4. EN PREPARACIÓN'],
        ['   • Preparando arreglos y decoración'],
        ['   • OPCIONAL: Descontar stock si se van a usar'],
        ['   • Insumos Descontados = TRUE'],
        [''],
        ['5. EN EVENTO'],
        ['   • Día del evento'],
        [''],
        ['6. FINALIZADO'],
        ['   • Evento terminado'],
        [''],
        ['7. RETIRADO'],
        ['   • Insumos fueron retirados del lugar'],
        ['   • ACCIÓN: Marcar qué se devolvió y qué falta'],
        ['   • Si hay faltantes: Insumos Faltantes = TRUE'],
        ['   • Sistema marca cliente con "insumo faltante en evento"'],
        [''],
        ['💡 CÁLCULO AUTOMÁTICO DE COSTOS'],
        [''],
        ['Costo Total = Costo Insumos + Costo Mano Obra + Costo Transporte + Costo Otros'],
        ['Precio Propuesta = Costo Total × (1 + Margen% / 100)'],
        ['Saldo = Precio Final - Anticipo'],
        [''],
        ['🔍 TIPOS DE INSUMO'],
        [''],
        ['• flor: Referencia a tabla Flores (ej: FL001)'],
        ['• contenedor: Referencia a tabla Contenedores (ej: CT001)'],
        ['• producto: Referencia a tabla Productos/Arreglos (ej: PR001)'],
        ['• producto_evento: Referencia a ProductoEvento (ej: PE001)'],
        ['• otro: Texto libre (mano de obra, transporte, otros servicios)'],
        [''],
        ['⚡ IMPORTACIÓN'],
        [''],
        ['Para importar estos datos al sistema:'],
        ['  1. Guardar este archivo como ESTRUCTURA_EVENTOS.xlsx'],
        ['  2. Colocar en carpeta backend/'],
        ['  3. Ejecutar: python3 importar_eventos_demo.py'],
        [''],
        ['El sistema importará:'],
        ['  • Productos de eventos (si no existen)'],
        ['  • Eventos completos'],
        ['  • Insumos asociados a cada evento'],
        [''],
        ['⚠️ ADVERTENCIAS EN LA INTERFAZ'],
        [''],
        ['Cuando un evento tiene "Insumos Faltantes = TRUE":'],
        ['  • Se muestra alerta visual (icono ⚠️ rojo)'],
        ['  • Tooltip con detalle de lo que falta'],
        ['  • El cliente queda marcado como "Insumo Faltante en Evento"'],
        ['  • Se bloquean nuevos eventos para ese cliente hasta devolución'],
        [''],
        ['📧 CONTACTO SOPORTE'],
        [''],
        ['Sistema desarrollado para Las Lira Florería'],
        ['Octubre 2025'],
    ]
    
    for row in instrucciones:
        ws_instrucciones.append(row)
    
    # Ajustar anchos
    ws_instrucciones.column_dimensions['A'].width = 100
    
    # Guardar archivo
    filename = 'ESTRUCTURA_EVENTOS.xlsx'
    wb.save(filename)
    print(f"✅ Archivo creado: {filename}")
    print(f"📊 Eventos: {len(eventos_data)}")
    print(f"📦 Insumos: {len(insumos_data)}")
    print(f"🎁 Productos: {len(productos_data)}")
    print()
    print("⚠️  EVENTO CON FALTANTES: EV004 (2 maceteros + 3 velas)")
    print()
    print("🚀 Para importar:")
    print("   python3 importar_eventos_demo.py")

if __name__ == "__main__":
    crear_excel_eventos()
