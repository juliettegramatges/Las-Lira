#!/usr/bin/env python3
"""
Genera un archivo Excel con la estructura del recetario de productos.
Muestra cómo se organizan los colores y flores para cada producto.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def crear_excel_recetario():
    """Crea un Excel con la estructura completa del recetario"""
    
    wb = openpyxl.Workbook()
    
    # ========================================
    # HOJA 1: PRODUCTOS CON COLORES
    # ========================================
    ws_productos = wb.active
    ws_productos.title = "Productos y Colores"
    
    # Headers
    headers_productos = [
        'ID Producto',
        'Nombre Producto',
        'Precio Venta',
        'ID Color',
        'Nombre Color',
        'Cantidad Tallos Sugerida',
        'Orden Color',
        'Notas Color'
    ]
    
    ws_productos.append(headers_productos)
    
    # Estilo para headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws_productos[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos de ejemplo
    datos_productos = [
        # Producto 1: Amor Eterno
        ['PR010', 'Amor Eterno', 65000, 1, 'Rojo', 12, 1, 'Color principal'],
        ['PR010', 'Amor Eterno', 65000, 2, 'Rosa', 8, 2, 'Color complementario'],
        
        # Producto 2: Campo Silvestre
        ['PR007', 'Campo Silvestre', 60000, 3, 'Morado', 4, 1, 'Flores moradas'],
        ['PR007', 'Campo Silvestre', 60000, 4, 'Verde/Follaje', 5, 2, 'Follaje y relleno'],
        
        # Producto 3: Dulce Lirio
        ['PR006', 'Dulce Lirio', 55000, 5, 'Rosado', 6, 1, 'Lirios rosados'],
        ['PR006', 'Dulce Lirio', 55000, 6, 'Blanco', 3, 2, 'Lirios blancos'],
    ]
    
    for fila in datos_productos:
        ws_productos.append(fila)
    
    # Ajustar anchos de columna
    ws_productos.column_dimensions['A'].width = 12
    ws_productos.column_dimensions['B'].width = 20
    ws_productos.column_dimensions['C'].width = 12
    ws_productos.column_dimensions['D'].width = 10
    ws_productos.column_dimensions['E'].width = 15
    ws_productos.column_dimensions['F'].width = 20
    ws_productos.column_dimensions['G'].width = 12
    ws_productos.column_dimensions['H'].width = 20
    
    # ========================================
    # HOJA 2: FLORES POR COLOR
    # ========================================
    ws_flores = wb.create_sheet("Flores por Color")
    
    headers_flores = [
        'ID Color',
        'Nombre Color',
        'ID Flor',
        'Nombre Flor',
        'Costo Unitario',
        'Es Predeterminada',
        'Stock Actual',
        'Stock En Uso',
        'Stock Disponible',
        'Notas'
    ]
    
    ws_flores.append(headers_flores)
    
    # Estilo headers
    for cell in ws_flores[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos de ejemplo
    datos_flores = [
        # Color Rojo (ID: 1)
        [1, 'Rojo', 'FL001', 'Rosa Roja', 1500, 'SÍ', 50, 12, 38, 'Predeterminada para rojo'],
        [1, 'Rojo', 'FL008', 'Clavel Rojo', 500, 'NO', 100, 0, 100, 'Alternativa económica'],
        
        # Color Rosa (ID: 2)
        [2, 'Rosa', 'FL002', 'Rosa Rosada', 1500, 'SÍ', 40, 8, 32, 'Predeterminada para rosa'],
        [2, 'Rosa', 'FL003', 'Tulipán Rosado', 2200, 'NO', 30, 0, 30, 'Opción premium'],
        
        # Color Morado (ID: 3)
        [3, 'Morado', 'FL015', 'Alstroemeria Morada', 1200, 'SÍ', 60, 4, 56, 'Predeterminada'],
        [3, 'Morado', 'FL020', 'Gerbera Morada', 1800, 'NO', 25, 0, 25, 'Alternativa llamativa'],
        
        # Color Verde/Follaje (ID: 4)
        [4, 'Verde/Follaje', 'FL025', 'Eucalipto Verde', 500, 'SÍ', 200, 5, 195, 'Follaje base'],
        [4, 'Verde/Follaje', 'FL021', 'Solidago', 800, 'NO', 80, 0, 80, 'Relleno amarillo'],
        
        # Color Rosado - Lirios (ID: 5)
        [5, 'Rosado', 'FL006', 'Lirio Rosado', 3000, 'SÍ', 20, 6, 14, 'Lirio principal'],
        
        # Color Blanco - Lirios (ID: 6)
        [6, 'Blanco', 'FL007', 'Lirio Blanco', 3000, 'SÍ', 25, 3, 22, 'Lirio complementario'],
    ]
    
    for fila in datos_flores:
        ws_flores.append(fila)
    
    # Ajustar anchos
    ws_flores.column_dimensions['A'].width = 10
    ws_flores.column_dimensions['B'].width = 15
    ws_flores.column_dimensions['C'].width = 10
    ws_flores.column_dimensions['D'].width = 18
    ws_flores.column_dimensions['E'].width = 14
    ws_flores.column_dimensions['F'].width = 16
    ws_flores.column_dimensions['G'].width = 12
    ws_flores.column_dimensions['H'].width = 12
    ws_flores.column_dimensions['I'].width = 15
    ws_flores.column_dimensions['J'].width = 25
    
    # Resaltar flores predeterminadas
    amarillo_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for row in ws_flores.iter_rows(min_row=2, max_row=ws_flores.max_row):
        if row[5].value == 'SÍ':  # Columna "Es Predeterminada"
            for cell in row:
                cell.fill = amarillo_fill
    
    # ========================================
    # HOJA 3: RESUMEN DE COSTOS POR PRODUCTO
    # ========================================
    ws_costos = wb.create_sheet("Costos por Producto")
    
    headers_costos = [
        'Producto',
        'Color',
        'Flor Predeterminada',
        'Cantidad',
        'Costo Unitario',
        'Costo Total Color',
        'Contenedor',
        'Costo Contenedor',
        'COSTO TOTAL',
        'PRECIO VENTA',
        'MARGEN',
        '% Margen'
    ]
    
    ws_costos.append(headers_costos)
    
    # Estilo headers
    for cell in ws_costos[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos de ejemplo
    datos_costos = [
        ['Amor Eterno', 'Rojo', 'Rosa Roja', 12, 1500, '=D2*E2', 'Macetero Cerámica', 8000, '=F2+H2', 65000, '=J2-I2', '=K2/J2'],
        ['Amor Eterno', 'Rosa', 'Rosa Rosada', 8, 1500, '=D3*E3', '', '', '=F3', '', '', ''],
        ['', '', '', '', 'TOTAL FLORES:', '=F2+F3', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', '', '', '', '', ''],
        ['Campo Silvestre', 'Morado', 'Alstroemeria Morada', 4, 1200, '=D6*E6', 'Canasto', 3000, '=F6+F7+H6', 60000, '=J6-I6', '=K6/J6'],
        ['Campo Silvestre', 'Verde/Follaje', 'Eucalipto Verde', 5, 500, '=D7*E7', '', '', '', '', '', ''],
        ['', '', '', '', 'TOTAL FLORES:', '=F6+F7', '', '', '', '', '', ''],
        ['', '', '', '', '', '', '', '', '', '', '', ''],
        ['Dulce Lirio', 'Rosado', 'Lirio Rosado', 6, 3000, '=D10*E10', 'Florero Vidrio', 5000, '=F10+F11+H10', 55000, '=J10-I10', '=K10/J10'],
        ['Dulce Lirio', 'Blanco', 'Lirio Blanco', 3, 3000, '=D11*E11', '', '', '', '', '', ''],
        ['', '', '', '', 'TOTAL FLORES:', '=F10+F11', '', '', '', '', '', ''],
    ]
    
    for fila in datos_costos:
        ws_costos.append(fila)
    
    # Formatear como moneda y porcentaje
    for row in ws_costos.iter_rows(min_row=2, max_row=ws_costos.max_row):
        # Costos (columnas E, F, H, I, J, K)
        for col_idx in [4, 5, 7, 8, 9, 10]:  # 0-indexed
            row[col_idx].number_format = '$#,##0'
        # Porcentaje (columna L)
        row[11].number_format = '0.0%'
    
    # Resaltar totales
    verde_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bold_font = Font(bold=True)
    
    for row_num in [2, 6, 10]:  # Filas de productos principales
        for cell in ws_costos[row_num]:
            cell.font = bold_font
    
    # Ajustar anchos
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws_costos.column_dimensions[col].width = 15
    
    # ========================================
    # HOJA 4: INSTRUCCIONES
    # ========================================
    ws_instrucciones = wb.create_sheet("📖 INSTRUCCIONES")
    
    instrucciones = [
        ['ESTRUCTURA DEL RECETARIO DE PRODUCTOS'],
        [''],
        ['Este archivo muestra cómo se organiza la información del recetario en la base de datos.'],
        [''],
        ['🌸 HOJA 1: Productos y Colores'],
        ['- Lista todos los productos con sus colores asociados'],
        ['- Cada producto puede tener múltiples colores (1, 2, 3...)'],
        ['- Cada color tiene una cantidad sugerida de tallos'],
        ['- El orden determina cómo se muestran en el simulador'],
        [''],
        ['🌺 HOJA 2: Flores por Color'],
        ['- Para cada color, muestra qué flores pueden usarse'],
        ['- La flor "Predeterminada" (SÍ) es la que se usa en el costo base'],
        ['- Puedes tener múltiples opciones de flores por color'],
        ['- Stock Disponible = Stock Actual - Stock En Uso'],
        [''],
        ['💰 HOJA 3: Costos por Producto'],
        ['- Calcula el costo total usando las flores predeterminadas'],
        ['- Suma el costo del contenedor'],
        ['- Muestra el margen de ganancia'],
        [''],
        ['⚙️ CÓMO SE GUARDA:'],
        ['- Los datos NO se guardan en Excel para el uso diario'],
        ['- Se guardan en la base de datos SQLite: backend/floreria.db'],
        ['- Excel solo se usa para importar datos iniciales'],
        ['- Cuando haces cambios en el simulador, se guardan en la DB'],
        [''],
        ['📊 FLUJO DE TRABAJO:'],
        ['1. Importación inicial: Excel → Script Python → Base de Datos'],
        ['2. Trabajo diario: Frontend ↔ API ↔ Base de Datos'],
        ['3. Modificar receta: Simulador → Endpoint API → Actualiza DB'],
        [''],
        ['🔄 PARA EXPORTAR DATOS ACTUALES:'],
        ['- Usa el script de exportación (próximamente)'],
        ['- O consulta directamente la base de datos'],
        [''],
        [f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'],
    ]
    
    for fila in instrucciones:
        ws_instrucciones.append(fila)
    
    # Estilo para título
    ws_instrucciones['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws_instrucciones.column_dimensions['A'].width = 80
    
    # Guardar
    filename = 'ESTRUCTURA_RECETARIO.xlsx'
    wb.save(filename)
    print(f"✅ Excel generado: {filename}")
    print(f"📂 Ubicación: {filename}")
    print(f"\n📋 Contenido:")
    print(f"   • Hoja 1: Productos y Colores")
    print(f"   • Hoja 2: Flores por Color (predeterminadas resaltadas)")
    print(f"   • Hoja 3: Costos por Producto (con cálculos)")
    print(f"   • Hoja 4: Instrucciones de uso")

if __name__ == "__main__":
    crear_excel_recetario()

