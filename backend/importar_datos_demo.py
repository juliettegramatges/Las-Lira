"""
Script para importar datos desde los archivos Excel demo a la base de datos
"""

import sys
import os
from datetime import datetime, date

# Configurar path
root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, root_dir)

from openpyxl import load_workbook
from backend.app import app, db
from backend.models.inventario import Flor, Contenedor, Bodega, Proveedor
from backend.models.producto import Producto, RecetaProducto
from backend.models.pedido import Pedido, PedidoInsumo
from backend.models.cliente import Cliente

def importar_proveedores():
    """Importar proveedores desde Excel"""
    print("\n📦 Importando proveedores...")
    wb = load_workbook('../06_Proveedores.xlsx')
    ws = wb.active
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Si no hay ID, saltar
            continue
            
        proveedor = Proveedor(
            id=row[0],
            nombre=row[1],
            contacto=row[2],
            telefono=row[3],
            email=row[4],
            especialidad=row[5],
            dias_entrega=row[6]
        )
        
        # Verificar si ya existe
        existing = Proveedor.query.get(proveedor.id)
        if not existing:
            db.session.add(proveedor)
            count += 1
    
    db.session.commit()
    print(f"✅ {count} proveedores importados")

def importar_bodegas():
    """Crear bodegas si no existen"""
    print("\n🏢 Creando bodegas...")
    
    # Verificar si ya existen bodegas
    if Bodega.query.count() > 0:
        print("✅ Bodegas ya existen")
        return
    
    bodegas = [
        Bodega(nombre='Bodega Principal', direccion='Almacén 1', encargado='Juan Pérez', telefono='+56912345678'),
        Bodega(nombre='Bodega Secundaria', direccion='Almacén 2', encargado='María González', telefono='+56912345679')
    ]
    
    for bodega in bodegas:
        db.session.add(bodega)
    
    db.session.commit()
    print(f"✅ {len(bodegas)} bodegas creadas")

def importar_flores():
    """Importar flores desde Excel"""
    print("\n🌸 Importando flores...")
    wb = load_workbook('../01_Inventario_Flores.xlsx')
    ws = wb.active
    
    # Crear mapeo de nombres de proveedores a IDs
    proveedores_map = {}
    for p in Proveedor.query.all():
        proveedores_map[p.nombre] = p.id
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        
        # Buscar proveedor_id por nombre
        proveedor_nombre = row[4] if row[4] else None
        proveedor_id = proveedores_map.get(proveedor_nombre) if proveedor_nombre else None
            
        flor = Flor(
            id=row[0],                          # ID
            tipo=row[1],                        # Tipo de Flor
            color=row[2],                       # Color
            foto_url=row[3] if row[3] else None, # Foto
            proveedor_id=proveedor_id,          # Proveedor (buscado por nombre)
            costo_unitario=float(row[5]) if row[5] else 0,  # Costo Unitario
            cantidad_stock=int(row[6]) if row[6] else 0,    # Cantidad Stock Actual
            unidad=row[7] if row[7] else 'tallo',           # Unidad
            fecha_actualizacion=datetime.strptime(row[8], '%Y-%m-%d').date() if row[8] else date.today()
        )
        
        existing = Flor.query.get(flor.id)
        if not existing:
            db.session.add(flor)
            count += 1
    
    db.session.commit()
    print(f"✅ {count} flores importadas")

def importar_contenedores():
    """Importar contenedores desde Excel"""
    print("\n🏺 Importando contenedores...")
    wb = load_workbook('../02_Inventario_Contenedores.xlsx')
    ws = wb.active
    
    # Obtener la primera bodega para asignar
    primera_bodega = Bodega.query.first()
    if not primera_bodega:
        print("⚠️  No hay bodegas disponibles")
        return
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        
        # Determinar bodega_id (1 o 2)
        bodega_nombre = row[9] if len(row) > 9 and row[9] else 'Bodega 1'
        bodega_id = 1 if '1' in bodega_nombre else 2
            
        contenedor = Contenedor(
            id=row[0],
            tipo=row[1],
            material=row[2],
            forma=row[3],
            tamano=row[4],
            color=row[5],
            foto_url=row[6] if row[6] else None,
            costo=float(row[7]) if row[7] else 0,
            stock=int(row[8]) if row[8] else 0,
            bodega_id=bodega_id,
            fecha_actualizacion=datetime.strptime(row[10], '%Y-%m-%d').date() if len(row) > 10 and row[10] else date.today()
        )
        
        existing = Contenedor.query.get(contenedor.id)
        if not existing:
            db.session.add(contenedor)
            count += 1
    
    db.session.commit()
    print(f"✅ {count} contenedores importados")

def importar_productos():
    """Importar productos desde Excel"""
    print("\n🎨 Importando productos...")
    wb = load_workbook('../03_Catalogo_Productos.xlsx')
    ws = wb.active
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        
        # Determinar tipo de arreglo basado en el contenedor
        tipos_macetero = row[5] if row[5] else ''
        if 'florero' in tipos_macetero.lower():
            tipo_arreglo = 'Con Florero'
        elif 'macetero' in tipos_macetero.lower():
            tipo_arreglo = 'Con Macetero'
        elif 'canasto' in tipos_macetero.lower():
            tipo_arreglo = 'Con Canasto'
        elif 'caja' in tipos_macetero.lower():
            tipo_arreglo = 'Con Caja'
        else:
            tipo_arreglo = 'Sin Contenedor'
            
        producto = Producto(
            id=row[0],
            nombre=row[1],
            descripcion=row[2],
            tipo_arreglo=tipo_arreglo,
            colores_asociados=row[3],
            flores_asociadas=row[4],
            tipos_macetero=row[5],
            vista_360_180=row[6],
            tamano=row[7],
            precio_venta=float(row[8]) if row[8] else 0,
            cuidados=row[9],
            imagen_url=row[10] if row[10] else None,
            disponible_shopify=(row[11] == 'Sí') if row[11] else True
        )
        
        try:
            db.session.add(producto)
            db.session.flush()  # Flush inmediato para detectar errores
            count += 1
        except Exception as e:
            print(f"⚠️  Error al importar producto {row[0]}: {e}")
            db.session.rollback()
            continue
    
    db.session.commit()
    print(f"✅ {count} productos importados")

def importar_recetas():
    """Importar recetas de productos desde Excel"""
    print("\n📝 Importando recetas de productos...")
    wb = load_workbook('../05_Recetas_Productos.xlsx')
    ws = wb.active
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        
        # Orden: ID Producto, Nombre Producto, ID Insumo, Tipo Insumo, Descripción, Cantidad, Unidad
        try:
            receta = RecetaProducto(
                producto_id=row[0],           # ID Producto
                insumo_id=row[2],             # ID Insumo
                insumo_tipo=row[3],           # Tipo Insumo (Flor/Contenedor)
                cantidad=int(row[5]) if row[5] else 1,  # Cantidad Necesaria
                unidad=row[6] if row[6] else 'unidad',  # Unidad
                es_opcional=False,
                notas=row[4] if row[4] else None  # Descripción
            )
            
            db.session.add(receta)
            count += 1
        except Exception as e:
            print(f"⚠️  Error en receta del producto {row[0]}: {e}")
            continue
    
    db.session.commit()
    print(f"✅ {count} recetas importadas")

def importar_pedidos():
    """Importar pedidos desde Excel"""
    print("\n🛍️ Importando pedidos...")
    wb = load_workbook('../04_Pedidos.xlsx')
    ws = wb.active
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        
        try:
            # Parsear fechas
            fecha_pedido = datetime.strptime(row[1], '%Y-%m-%d %H:%M') if row[1] else datetime.now()
            fecha_entrega = datetime.strptime(row[2], '%Y-%m-%d %H:%M') if row[2] else datetime.now()
            
            pedido = Pedido(
                id=row[0],
                fecha_pedido=fecha_pedido,
                fecha_entrega=fecha_entrega,
                dia_entrega=row[3],
                canal=row[4],
                shopify_order_number=row[5] if row[5] else None,
                cliente_nombre=row[6],
                cliente_telefono=row[7],
                arreglo_pedido=row[8],
                detalles_adicionales=row[9] if row[9] else None,
                precio_ramo=float(row[10]) if row[10] else 0,
                precio_envio=float(row[11]) if row[11] else 0,
                destinatario=row[12] if row[12] else None,
                mensaje=row[13] if row[13] else None,
                firma=row[14] if row[14] else None,
                direccion_entrega=row[15],
                comuna=row[16] if row[16] else None,
                motivo=row[17] if row[17] else None,
                estado=row[18] if row[18] else 'Pedido',
                estado_pago=row[19] if row[19] else 'No Pagado',
                tipo_pedido=row[20] if row[20] else None,
                cobranza=row[21] if row[21] else None,
                foto_enviado_url=row[22] if row[22] else None
            )
            
            existing = Pedido.query.get(pedido.id)
            if not existing:
                db.session.add(pedido)
                count += 1
        except Exception as e:
            print(f"⚠️  Error en pedido {row[0]}: {e}")
            continue
    
    db.session.commit()
    print(f"✅ {count} pedidos importados")

def importar_clientes():
    """Importar clientes desde Excel (sin estadísticas, se calculan después)"""
    print("\n👥 Importando clientes...")
    wb = load_workbook('../07_Clientes.xlsx')
    ws = wb['Clientes']
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Si no hay ID, saltar
            continue
        
        try:
            cliente = Cliente(
                id=row[0],
                nombre=row[1],
                telefono=row[2],
                email=row[3] if row[3] else None,
                tipo_cliente=row[4] if row[4] else 'Nuevo',
                direccion_principal=row[5] if row[5] else None,
                notas=row[6] if row[6] else None,
                total_pedidos=0,  # Se calculará después
                total_gastado=0   # Se calculará después
            )
            
            existing = Cliente.query.get(cliente.id)
            if not existing:
                db.session.add(cliente)
                count += 1
        except Exception as e:
            print(f"⚠️ Error al importar cliente {row[0]}: {e}")
            continue
    
    db.session.commit()
    print(f"✅ {count} clientes importados (estadísticas se calcularán desde pedidos)")

def vincular_pedidos_a_clientes():
    """Vincular pedidos existentes a clientes y calcular estadísticas"""
    print("\n🔗 Vinculando pedidos a clientes y calculando estadísticas...")
    
    # Normalizar teléfono (quitar espacios, guiones, paréntesis)
    import re
    def normalizar_telefono(telefono):
        return re.sub(r'[\s\-\(\)]', '', telefono)
    
    # Obtener todos los pedidos
    pedidos = Pedido.query.all()
    clientes_actualizados = set()
    
    # Vincular cada pedido a su cliente
    for pedido in pedidos:
        if not pedido.cliente_telefono:
            continue
        
        # Buscar cliente por teléfono normalizado
        telefono_pedido = normalizar_telefono(pedido.cliente_telefono)
        cliente = None
        
        for c in Cliente.query.all():
            if normalizar_telefono(c.telefono) == telefono_pedido:
                cliente = c
                break
        
        if cliente:
            # Vincular pedido al cliente
            pedido.cliente_id = cliente.id
            clientes_actualizados.add(cliente.id)
        else:
            print(f"⚠️  Cliente no encontrado para pedido {pedido.id} (tel: {pedido.cliente_telefono})")
    
    db.session.commit()
    print(f"✅ Pedidos vinculados a clientes")
    
    # Recalcular estadísticas de cada cliente
    print("\n📊 Calculando estadísticas de clientes desde pedidos reales...")
    for cliente_id in clientes_actualizados:
        cliente = Cliente.query.get(cliente_id)
        if not cliente:
            continue
        
        # Obtener todos los pedidos del cliente
        pedidos_cliente = Pedido.query.filter_by(cliente_id=cliente.id).all()
        
        # Calcular estadísticas
        cliente.total_pedidos = len(pedidos_cliente)
        cliente.total_gastado = sum((p.precio_ramo or 0) + (p.precio_envio or 0) for p in pedidos_cliente)
        
        # Actualizar última compra
        if pedidos_cliente:
            pedidos_ordenados = sorted(pedidos_cliente, key=lambda p: p.fecha_pedido, reverse=True)
            cliente.ultima_compra = pedidos_ordenados[0].fecha_pedido
        
        print(f"  ✓ {cliente.nombre}: {cliente.total_pedidos} pedidos, ${cliente.total_gastado:,.0f}")
    
    db.session.commit()
    print(f"✅ Estadísticas calculadas para {len(clientes_actualizados)} clientes")

def main():
    """Ejecutar todas las importaciones"""
    print("=" * 60)
    print("🌸 IMPORTACIÓN DE DATOS DEMO - LAS LIRA 🌸")
    print("=" * 60)
    
    with app.app_context():
        # Crear tablas si no existen
        db.create_all()
        print("✅ Tablas de base de datos verificadas")
        
        try:
            # Importar en orden de dependencias
            importar_bodegas()
            importar_proveedores()
            importar_flores()
            importar_contenedores()
            importar_productos()
            importar_recetas()
            importar_clientes()  # Importar clientes ANTES de pedidos
            importar_pedidos()
            vincular_pedidos_a_clientes()  # Vincular y calcular estadísticas
            
            print("\n" + "=" * 60)
            print("✨ ¡IMPORTACIÓN COMPLETADA EXITOSAMENTE! ✨")
            print("=" * 60)
            print("\n🌐 Recarga la página web para ver los datos")
            print("📊 Revisa las siguientes secciones:")
            print("   • Inventario: Flores y Contenedores")
            print("   • Productos: Catálogo de arreglos")
            print("   • Clientes: Base de datos de clientes")
            print("   • Pedidos: Órdenes de clientes")
            print("   • Tablero: Vista Kanban")
            print("   • Rutas: Optimización por comuna")
            
        except Exception as e:
            print(f"\n❌ Error durante la importación: {e}")
            import traceback
            traceback.print_exc()
            db.session.rollback()

if __name__ == '__main__':
    main()

