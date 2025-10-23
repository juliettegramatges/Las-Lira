#!/usr/bin/env python3
"""
Importa eventos desde Excel a la base de datos
"""

import sys
import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app, db
from models.evento import Evento, EventoInsumo, ProductoEvento
from models.inventario import Flor, Contenedor
from models.producto import Producto
from models.cliente import Cliente
import openpyxl
from datetime import datetime

def importar_productos_evento():
    """Importa productos específicos de eventos"""
    print("📦 Importando productos de eventos...")
    
    try:
        wb = openpyxl.load_workbook('ESTRUCTURA_EVENTOS.xlsx')
        ws = wb['03_Productos_Evento']
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Si no hay código, saltar
                continue
            
            codigo, nombre, categoria, stock, en_evento, _, costo_compra, costo_alquiler, \
            descripcion, medidas, color, material = row
            
            # Verificar si ya existe
            existe = ProductoEvento.query.filter_by(codigo=codigo).first()
            if existe:
                print(f"  ⏭️  {codigo} ya existe, actualizando...")
                existe.nombre = nombre
                existe.categoria = categoria
                existe.cantidad_stock = stock or 0
                existe.cantidad_en_evento = en_evento or 0
                existe.costo_compra = costo_compra or 0
                existe.costo_alquiler = costo_alquiler or 0
                existe.descripcion = descripcion
                existe.medidas = medidas
                existe.color = color
                existe.material = material
            else:
                nuevo = ProductoEvento(
                    codigo=codigo,
                    nombre=nombre,
                    categoria=categoria,
                    cantidad_stock=stock or 0,
                    cantidad_en_evento=en_evento or 0,
                    costo_compra=costo_compra or 0,
                    costo_alquiler=costo_alquiler or 0,
                    descripcion=descripcion,
                    medidas=medidas,
                    color=color,
                    material=material
                )
                db.session.add(nuevo)
                print(f"  ✅ {codigo} - {nombre}")
        
        db.session.commit()
        print(f"✨ Productos de eventos importados!\n")
        
    except FileNotFoundError:
        print("⚠️  ESTRUCTURA_EVENTOS.xlsx no encontrado")
    except Exception as e:
        print(f"❌ Error: {e}")
        db.session.rollback()


def importar_eventos():
    """Importa eventos desde Excel"""
    print("📅 Importando eventos...")
    
    try:
        wb = openpyxl.load_workbook('ESTRUCTURA_EVENTOS.xlsx')
        ws = wb['01_Eventos']
        
        eventos_importados = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Si no hay ID, saltar
                continue
            
            evento_id = row[0]
            
            # Verificar si ya existe
            existe = Evento.query.get(evento_id)
            if existe:
                print(f"  ⏭️  {evento_id} ya existe, saltando...")
                continue
            
            # Parsear fecha
            fecha_evento = None
            if row[6]:  # Fecha Evento
                if isinstance(row[6], datetime):
                    fecha_evento = row[6]
                else:
                    try:
                        fecha_evento = datetime.strptime(str(row[6]), '%Y-%m-%d')
                    except:
                        pass
            
            # Parsear fecha cotización
            fecha_cotizacion = None
            if row[26]:  # Fecha Cotización
                if isinstance(row[26], datetime):
                    fecha_cotizacion = row[26]
                else:
                    try:
                        fecha_cotizacion = datetime.strptime(str(row[26]), '%Y-%m-%d')
                    except:
                        fecha_cotizacion = datetime.utcnow()
            else:
                fecha_cotizacion = datetime.utcnow()
            
            nuevo_evento = Evento(
                id=evento_id,
                cliente_nombre=row[1],
                cliente_telefono=row[2],
                cliente_email=row[3],
                nombre_evento=row[4],
                tipo_evento=row[5],
                fecha_evento=fecha_evento,
                hora_evento=row[7],
                lugar_evento=row[8],
                cantidad_personas=row[9] or 0,
                estado=row[10] or 'Cotización',
                costo_insumos=row[11] or 0,
                costo_mano_obra=row[12] or 0,
                costo_transporte=row[13] or 0,
                costo_otros=row[14] or 0,
                costo_total=row[15] or 0,
                margen_porcentaje=row[16] or 30,
                precio_propuesta=row[17] or 0,
                precio_final=row[18] or 0,
                anticipo=row[19] or 0,
                saldo=row[20] or 0,
                pagado=row[21] or False,
                insumos_reservados=row[22] or False,
                insumos_descontados=row[23] or False,
                insumos_faltantes=row[24] or False,
                notas_cotizacion=row[25],
                fecha_cotizacion=fecha_cotizacion
            )
            
            db.session.add(nuevo_evento)
            eventos_importados += 1
            print(f"  ✅ {evento_id} - {nuevo_evento.nombre_evento}")
        
        db.session.commit()
        print(f"✨ {eventos_importados} eventos importados!\n")
        
    except FileNotFoundError:
        print("⚠️  ESTRUCTURA_EVENTOS.xlsx no encontrado")
    except Exception as e:
        print(f"❌ Error: {e}")
        db.session.rollback()


def importar_insumos_eventos():
    """Importa insumos de eventos desde Excel"""
    print("📦 Importando insumos de eventos...")
    
    try:
        wb = openpyxl.load_workbook('ESTRUCTURA_EVENTOS.xlsx')
        ws = wb['02_Insumos_Eventos']
        
        insumos_importados = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Si no hay evento_id, saltar
                continue
            
            evento_id = row[0]
            tipo_insumo = row[1]
            id_insumo = row[2]
            nombre = row[3]
            cantidad = row[4] or 1
            costo_unitario = row[5] or 0
            costo_total = row[6] or (cantidad * costo_unitario)
            reservado = row[7] or False
            descontado = row[8] or False
            devuelto = row[9] or False
            cantidad_faltante = row[10] or 0
            notas = row[11]
            
            # Verificar que el evento existe
            evento = Evento.query.get(evento_id)
            if not evento:
                print(f"  ⚠️  Evento {evento_id} no existe, saltando insumo...")
                continue
            
            nuevo_insumo = EventoInsumo(
                evento_id=evento_id,
                tipo_insumo=tipo_insumo,
                cantidad=cantidad,
                costo_unitario=costo_unitario,
                costo_total=costo_total,
                reservado=reservado,
                descontado_stock=descontado,
                devuelto=devuelto,
                cantidad_faltante=cantidad_faltante,
                notas=notas
            )
            
            # Asignar referencia según tipo
            if tipo_insumo == 'producto':
                nuevo_insumo.producto_id = id_insumo
            elif tipo_insumo == 'producto_evento':
                # Buscar por código
                prod_evento = ProductoEvento.query.filter_by(codigo=id_insumo).first()
                if prod_evento:
                    nuevo_insumo.producto_evento_id = prod_evento.id
            elif tipo_insumo == 'contenedor':
                nuevo_insumo.contenedor_id = id_insumo
            elif tipo_insumo == 'flor':
                nuevo_insumo.flor_id = id_insumo
            elif tipo_insumo == 'otro':
                nuevo_insumo.nombre_otro = nombre
            
            db.session.add(nuevo_insumo)
            insumos_importados += 1
            print(f"  ✅ {evento_id} - {tipo_insumo}: {nombre}")
        
        db.session.commit()
        print(f"✨ {insumos_importados} insumos importados!\n")
        
    except FileNotFoundError:
        print("⚠️  ESTRUCTURA_EVENTOS.xlsx no encontrado")
    except Exception as e:
        print(f"❌ Error importando insumos: {e}")
        import traceback
        traceback.print_exc()
        db.session.rollback()


def main():
    """Ejecuta la importación completa"""
    with app.app_context():
        print("\n" + "="*60)
        print("🌸 IMPORTACIÓN DE EVENTOS - LAS LIRA")
        print("="*60 + "\n")
        
        # 1. Importar productos de eventos primero (para las referencias)
        importar_productos_evento()
        
        # 2. Importar eventos
        importar_eventos()
        
        # 3. Importar insumos de eventos
        importar_insumos_eventos()
        
        print("="*60)
        print("✅ IMPORTACIÓN COMPLETADA")
        print("="*60 + "\n")
        
        # Mostrar resumen
        total_eventos = Evento.query.count()
        total_productos_evento = ProductoEvento.query.count()
        total_insumos = EventoInsumo.query.count()
        
        print(f"📊 RESUMEN:")
        print(f"   • Eventos en BD: {total_eventos}")
        print(f"   • Productos de Eventos: {total_productos_evento}")
        print(f"   • Insumos de Eventos: {total_insumos}")
        print()


if __name__ == "__main__":
    main()

