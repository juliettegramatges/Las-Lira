"""
Utilidades para generación y estilizado de archivos Excel
"""

from openpyxl.styles import Font, PatternFill, Alignment


def aplicar_estilo_encabezado(worksheet, headers, row=1):
    """
    Aplica estilo consistente a los encabezados de una hoja Excel

    Args:
        worksheet: Hoja de trabajo de openpyxl
        headers (list): Lista de nombres de encabezados
        row (int): Número de fila donde aplicar encabezados (default: 1)

    Returns:
        worksheet: La hoja con los encabezados estilizados
    """
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    return worksheet


def ajustar_ancho_columnas(worksheet, min_width=10, max_width=50):
    """
    Ajusta automáticamente el ancho de las columnas según el contenido

    Args:
        worksheet: Hoja de trabajo de openpyxl
        min_width (int): Ancho mínimo de columna
        max_width (int): Ancho máximo de columna

    Returns:
        worksheet: La hoja con columnas ajustadas
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    return worksheet


def crear_workbook_con_encabezado(titulo, headers):
    """
    Crea un workbook nuevo con encabezados estilizados

    Args:
        titulo (str): Título de la hoja
        headers (list): Lista de nombres de encabezados

    Returns:
        tuple: (workbook, worksheet) listos para agregar datos
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = titulo

    aplicar_estilo_encabezado(ws, headers)

    return wb, ws
