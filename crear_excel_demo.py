#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para crear archivos Excel demo para el sistema de gestión de florería Las-Lira
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

def crear_estilo_header(ws, row=1):
    """Aplica estilo a la fila de encabezado"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def crear_inventario_flores():
    """Crea archivo de inventario de flores (NO tienen bodega, se compran según necesidad)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Flores"
    
    # Encabezados
    headers = ["ID", "Tipo de Flor", "Color", "Foto", "Proveedor", "Costo Unitario", "Cantidad Stock Actual", "Unidad", "Última Actualización"]
    ws.append(headers)
    crear_estilo_header(ws)
    
    # Datos demo
    flores = [
        ["FL001", "Rosa", "Rojo", "rosa-roja.jpg", "Flores del Valle", 1500, 120, "Tallo", "2025-10-20"],
        ["FL002", "Rosa", "Blanco", "rosa-blanca.jpg", "Flores del Valle", 1500, 85, "Tallo", "2025-10-20"],
        ["FL003", "Rosa", "Rosado", "rosa-rosada.jpg", "Flores del Valle", 1500, 95, "Tallo", "2025-10-20"],
        ["FL004", "Rosa", "Amarillo", "rosa-amarilla.jpg", "Flores del Valle", 1500, 60, "Tallo", "2025-10-21"],
        ["FL005", "Lirio", "Blanco", "lirio-blanco.jpg", "Jardín Central", 2500, 45, "Tallo", "2025-10-19"],
        ["FL006", "Lirio", "Rosado", "lirio-rosado.jpg", "Jardín Central", 2500, 38, "Tallo", "2025-10-19"],
        ["FL007", "Girasol", "Amarillo", "girasol.jpg", "Campo Florido", 2000, 30, "Tallo", "2025-10-22"],
        ["FL008", "Clavel", "Rojo", "clavel-rojo.jpg", "Flores del Valle", 800, 150, "Tallo", "2025-10-20"],
        ["FL009", "Clavel", "Blanco", "clavel-blanco.jpg", "Flores del Valle", 800, 140, "Tallo", "2025-10-20"],
        ["FL010", "Clavel", "Rosado", "clavel-rosado.jpg", "Flores del Valle", 800, 130, "Tallo", "2025-10-20"],
        ["FL011", "Tulipán", "Rojo", "tulipan-rojo.jpg", "Jardín Central", 2200, 25, "Tallo", "2025-10-18"],
        ["FL012", "Tulipán", "Amarillo", "tulipan-amarillo.jpg", "Jardín Central", 2200, 20, "Tallo", "2025-10-18"],
        ["FL013", "Hortensia", "Azul", "hortensia-azul.jpg", "Campo Florido", 3500, 15, "Ramo", "2025-10-21"],
        ["FL014", "Hortensia", "Rosado", "hortensia-rosada.jpg", "Campo Florido", 3500, 18, "Ramo", "2025-10-21"],
        ["FL015", "Gerbera", "Naranja", "gerbera-naranja.jpg", "Flores del Valle", 1800, 55, "Tallo", "2025-10-20"],
        ["FL016", "Gerbera", "Rosado", "gerbera-rosada.jpg", "Flores del Valle", 1800, 48, "Tallo", "2025-10-20"],
        ["FL017", "Orquídea", "Blanco", "orquidea-blanca.jpg", "Jardín Central", 5000, 12, "Tallo", "2025-10-19"],
        ["FL018", "Alstroemeria", "Morado", "alstroemeria-morada.jpg", "Campo Florido", 1200, 70, "Tallo", "2025-10-22"],
        ["FL019", "Eucalipto", "Verde", "eucalipto.jpg", "Campo Florido", 500, 200, "Rama", "2025-10-21"],
        ["FL020", "Solidago", "Amarillo", "solidago.jpg", "Flores del Valle", 600, 100, "Rama", "2025-10-20"],
    ]
    
    for fila in flores:
        ws.append(fila)
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 18
    
    wb.save("01_Inventario_Flores.xlsx")
    print("✅ Archivo '01_Inventario_Flores.xlsx' creado")

def crear_inventario_contenedores():
    """Crea archivo de inventario de contenedores (floreros, maceteros, canastos)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contenedores"
    
    # Encabezados
    headers = ["ID", "Tipo", "Material", "Forma", "Tamaño", "Color", "Foto", "Costo", "Stock", "Bodega", "Última Actualización"]
    ws.append(headers)
    crear_estilo_header(ws)
    
    # Datos demo
    contenedores = [
        ["CO001", "Florero", "Vidrio", "Cilíndrico", "Grande (30cm)", "Transparente", "florero-vidrio-grande.jpg", 3500, 15, "Bodega 1", "2025-10-15"],
        ["CO002", "Florero", "Vidrio", "Cilíndrico", "Mediano (20cm)", "Transparente", "florero-vidrio-mediano.jpg", 2500, 22, "Bodega 1", "2025-10-15"],
        ["CO003", "Florero", "Vidrio", "Cilíndrico", "Pequeño (15cm)", "Transparente", "florero-vidrio-pequeno.jpg", 1500, 30, "Bodega 2", "2025-10-15"],
        ["CO004", "Florero", "Cerámica", "Redondo", "Mediano (18cm)", "Blanco", "florero-ceramica-blanco.jpg", 4000, 12, "Bodega 1", "2025-10-18"],
        ["CO005", "Florero", "Cerámica", "Cuadrado", "Pequeño (12cm)", "Negro", "florero-ceramica-negro.jpg", 3000, 18, "Bodega 1", "2025-10-18"],
        ["CO006", "Macetero", "Cerámica", "Redondo", "Grande (25cm)", "Terracota", "macetero-terracota-grande.jpg", 5000, 10, "Bodega 2", "2025-10-19"],
        ["CO007", "Macetero", "Cerámica", "Redondo", "Mediano (18cm)", "Terracota", "macetero-terracota-mediano.jpg", 3500, 15, "Bodega 2", "2025-10-19"],
        ["CO008", "Macetero", "Cerámica", "Redondo", "Pequeño (12cm)", "Terracota", "macetero-terracota-pequeno.jpg", 2000, 25, "Bodega 2", "2025-10-19"],
        ["CO009", "Macetero", "Plástico", "Redondo", "Mediano (18cm)", "Blanco", "macetero-plastico-blanco.jpg", 1500, 30, "Bodega 1", "2025-10-20"],
        ["CO010", "Canasto", "Mimbre", "Rectangular", "Grande (35cm)", "Natural", "canasto-rectangular-grande.jpg", 4500, 8, "Bodega 1", "2025-10-16"],
        ["CO011", "Canasto", "Mimbre", "Rectangular", "Mediano (25cm)", "Natural", "canasto-rectangular-mediano.jpg", 3000, 12, "Bodega 1", "2025-10-16"],
        ["CO012", "Canasto", "Mimbre", "Redondo", "Pequeño (20cm)", "Natural", "canasto-redondo-pequeno.jpg", 2500, 15, "Bodega 2", "2025-10-16"],
        ["CO013", "Florero", "Vidrio", "Burbuja", "Pequeño (10cm)", "Transparente", "florero-burbuja.jpg", 2000, 20, "Bodega 1", "2025-10-17"],
        ["CO014", "Macetero", "Cerámica", "Cuadrado", "Grande (30cm)", "Gris", "macetero-ceramica-gris.jpg", 6000, 6, "Bodega 2", "2025-10-21"],
        ["CO015", "Canasto", "Mimbre", "Ovalado", "Mediano (28cm)", "Natural", "canasto-ovalado.jpg", 3500, 10, "Bodega 1", "2025-10-16"],
    ]
    
    for fila in contenedores:
        ws.append(fila)
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 18
    
    wb.save("02_Inventario_Contenedores.xlsx")
    print("✅ Archivo '02_Inventario_Contenedores.xlsx' creado")

def crear_productos_catalogo():
    """Crea archivo con catálogo de productos (arreglos predefinidos)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo Productos"
    
    # Encabezados
    headers = ["ID", "Nombre Producto", "Descripción", "Colores Asociados", "Flores Asociadas", 
               "Tipos Macetero Posibles", "Vista (360/180)", "Tamaño (cm)", "Precio Venta", 
               "Cuidados", "Foto", "Disponible Shopify"]
    ws.append(headers)
    crear_estilo_header(ws)
    
    # Productos demo con información más detallada
    productos = [
        ["PR001", "Pasión Roja", "Arreglo elegante en tonos rojos", "Rojo, Verde oscuro, Burdeo",
         "Rosa roja, Clavel rojo, Eucalipto", "Florero vidrio cilíndrico", "360", "25 x 35", 35000,
         "Cambiar agua cada 2 días, evitar luz directa, cortar tallos en diagonal", "passion-roja.jpg", "Sí"],
        
        ["PR002", "Sueño Blanco", "Delicado arreglo en blancos puros", "Blanco, Verde claro",
         "Rosa blanca, Lirio blanco, Gerbera blanca", "Florero vidrio redondo", "360", "22 x 30", 32000,
         "Cambiar agua diariamente, mantener en lugar fresco, eliminar polen de lirios", "sueno-blanco.jpg", "Sí"],
        
        ["PR003", "Jardín Primaveral", "Mezcla de colores vibrantes", "Amarillo, Naranja, Rosado, Morado",
         "Gerbera, Alstroemeria, Rosa, Solidago", "Florero vidrio grande", "360", "30 x 40", 42000,
         "Cambiar agua cada 2-3 días, exposición luz indirecta, agregar nutriente floral", "jardin-primaveral.jpg", "Sí"],
        
        ["PR004", "Elegancia Rosa", "Rosas rosadas en florero", "Rosado, Verde, Blanco",
         "Rosa rosada, Eucalipto, Solidago", "Florero cerámica blanco", "180", "20 x 35", 38000,
         "Cambiar agua cada 2 días, cortar 1cm de tallo cada 3 días, ambiente fresco", "elegancia-rosa.jpg", "Sí"],
        
        ["PR005", "Sol Radiante", "Girasoles y flores amarillas", "Amarillo, Naranja, Verde",
         "Girasol, Gerbera naranja, Solidago", "Florero vidrio cilíndrico", "360", "28 x 38", 30000,
         "Cambiar agua diariamente, requiere buena luz, girasoles duran 7-10 días", "sol-radiante.jpg", "Sí"],
        
        ["PR006", "Dulce Lirio", "Lirios blancos y rosados", "Blanco, Rosado, Verde",
         "Lirio blanco, Lirio rosado, Eucalipto", "Florero vidrio burbuja", "360", "25 x 40", 45000,
         "Cambiar agua cada 2 días, quitar polen para evitar manchas, cortar en diagonal", "dulce-lirio.jpg", "Sí"],
        
        ["PR007", "Campo Silvestre", "Arreglo rústico en canasto", "Multicolor natural",
         "Mix de temporada, Gerbera, Alstroemeria, Follaje", "Canasto mimbre rectangular", "180", "35 x 25", 48000,
         "Verificar nivel de agua en esponja, rociar flores 1 vez al día", "campo-silvestre.jpg", "Sí"],
        
        ["PR008", "Orquídea Imperial", "Orquídeas blancas premium", "Blanco puro",
         "Orquídea phalaenopsis", "Macetero cerámica gris", "360", "18 x 45", 55000,
         "Regar 1 vez por semana, luz indirecta brillante, no mojar flores", "orquidea-imperial.jpg", "Sí"],
        
        ["PR009", "Ramo Clásico", "Ramo de rosas rojas", "Rojo intenso, Verde",
         "Rosa roja, Eucalipto", "Sin contenedor (ramo)", "360", "Ø 25", 28000,
         "Cortar tallos y colocar en agua inmediatamente, cambiar agua diariamente", "ramo-clasico.jpg", "Sí"],
        
        ["PR010", "Amor Eterno", "Caja con rosas", "Rojo, Rosado suave",
         "Rosa roja, Rosa rosada", "Caja cuadrada", "360", "25 x 25", 65000,
         "Rociar con agua 1-2 veces al día, no exponer al sol directo", "amor-eterno.jpg", "Sí"],
    ]
    
    for fila in productos:
        ws.append(fila)
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 50
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 15
    
    wb.save("03_Catalogo_Productos.xlsx")
    print("✅ Archivo '03_Catalogo_Productos.xlsx' creado")

def crear_pedidos():
    """Crea archivo de pedidos con todos los campos necesarios"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    
    # Encabezados
    headers = ["ID Pedido", "Fecha Pedido", "Fecha Entrega", "Día Entrega", "Canal", "Nro Pedido Shopify", 
               "Nombre Cliente", "Celular", "Arreglo Pedido", "Detalles Adicionales", 
               "Precio Ramo", "Precio Envío", "Para (Destinatario)", "Mensaje", "Firma", 
               "Dirección", "Comuna", "Motivo", "Estado", "Estado Pago", "Tipo Pedido", "Cobranza", "Foto Enviado"]
    ws.append(headers)
    crear_estilo_header(ws)
    
    # Datos demo de pedidos SINCRONIZADOS con clientes existentes
    pedidos = [
        # María González (CLI001, +56912345678) - Cliente Fiel - 8 pedidos = $280.000
        ["PED001", "2025-10-20 09:30", "2025-10-21 14:00", "LUNES", "Shopify", "#SH1234", 
         "María González", "+56912345678", "Pasión Roja", "Sin cambios", 
         33000, 7000, "Ana González", "Feliz cumpleaños hermana", "Con cariño, María",
         "Av. Apoquindo 1234, Las Condes", "Las Condes", "Cumpleaños", "Despachados", "Pagado", "Normal", "BOLETA 11248 TR. 21/10/25", ""],
        
        ["PED002", "2025-09-15 10:00", "2025-09-16 15:00", "LUNES", "WhatsApp", "", 
         "María González", "+56912345678", "Rosas Rojas Premium", "", 
         29000, 7000, "Laura González", "Te quiero mucho mamá", "María",
         "Av. Apoquindo 1234, Las Condes", "Las Condes", "Día de la Madre", "Despachados", "Pagado", "Normal", "", ""],
        
        ["PED003", "2025-08-20 14:30", "2025-08-21 11:00", "MARTES", "Shopify", "#SH1100", 
         "María González", "+56912345678", "Elegancia Rosa", "", 
         28000, 7000, "Carlos González", "Feliz aniversario", "Tu esposa",
         "Av. Apoquindo 1234, Las Condes", "Las Condes", "Aniversario", "Archivado", "Pagado", "Normal", "", ""],
        
        ["PED016", "2025-07-10 11:00", "2025-07-11 15:00", "JUEVES", "WhatsApp", "", 
         "María González", "+56912345678", "Rosas Rojas Clásicas", "Flores favoritas", 
         28000, 7000, "Patricia González", "Te extraño", "María",
         "Av. Apoquindo 1234, Las Condes", "Las Condes", "Sin motivo", "Archivado", "Pagado", "Normal", "", ""],
        
        ["PED017", "2025-06-05 09:30", "2025-06-06 10:00", "MIERCOLES", "Shopify", "#SH1050", 
         "María González", "+56912345678", "Jardín de Rosas", "", 
         28000, 7000, "Elena González", "Felicidades mamá", "Tu hija",
         "Av. Apoquindo 1234, Las Condes", "Las Condes", "Día de la Madre", "Archivado", "Pagado", "Normal", "", ""],
        
        ["PED018", "2025-05-15 14:00", "2025-05-16 11:00", "JUEVES", "WhatsApp", "", 
         "María González", "+56912345678", "Bouquet Romántico", "Con tarjeta", 
         25000, 7000, "Roberto González", "Te amo", "María",
         "Av. Apoquindo 1234, Las Condes", "Las Condes", "San Valentín", "Archivado", "Pagado", "Normal", "", ""],
        
        ["PED019", "2025-04-20 10:30", "2025-04-21 14:00", "SABADO", "Shopify", "#SH1020", 
         "María González", "+56912345678", "Rosas Premium Mix", "", 
         28000, 7000, "Sofía González", "Feliz cumpleaños", "María",
         "Av. Apoquindo 1234, Las Condes", "Las Condes", "Cumpleaños", "Archivado", "Pagado", "Normal", "", ""],
        
        ["PED020", "2025-03-12 16:00", "2025-03-13 09:00", "MIERCOLES", "WhatsApp", "", 
         "María González", "+56912345678", "Arreglo Elegante", "Rosas rojas grandes", 
         25000, 7000, "Andrés González", "Gracias por todo", "María",
         "Av. Apoquindo 1234, Las Condes", "Las Condes", "Agradecimiento", "Archivado", "Pagado", "Normal", "", ""],
        
        # Juan Pérez (CLI002, +56987654321) - Cliente VIP - 15 pedidos (mostramos 3)
        ["PED004", "2025-10-22 14:20", "2025-10-23 11:00", "MIERCOLES", "Shopify", "#SH1236", 
         "Juan Pérez", "+56987654321", "Arreglo Corporativo XL", "Para oficina principal", 
         85000, 10000, "", "", "",
         "Av. Vitacura 5678, Vitacura", "Vitacura", "Corporativo", "En Proceso", "Pagado", "EVENTO", "", ""],
        
        ["PED005", "2025-10-10 09:00", "2025-10-11 08:00", "VIERNES", "Shopify", "#SH1200", 
         "Juan Pérez", "+56987654321", "Centro de Mesa Premium", "Evento ejecutivo", 
         95000, 10000, "", "", "",
         "Av. Vitacura 5678, Vitacura", "Vitacura", "Evento", "Archivado", "Pagado", "EVENTO", "", ""],
        
        # Ana Martínez (CLI003, +56923456789) - Cumplidor - 5 pedidos
        ["PED006", "2025-10-22 08:45", "2025-10-23 16:00", "MIERCOLES", "Shopify", "#SH1235", 
         "Ana Martínez", "+56923456789", "Jardín Primaveral", "", 
         42000, 7000, "Claudia Ramírez", "Que te mejores pronto", "Tu amiga Ana",
         "Los Militares 2345, Las Condes", "Las Condes", "Mejórate", "Listo para Despacho", "Pagado", "Normal", "BOLETA 11249 TR. 23/10/25", ""],
        
        ["PED007", "2025-09-05 11:00", "2025-09-06 14:00", "JUEVES", "WhatsApp", "", 
         "Ana Martínez", "+56923456789", "Ramo de Lirios", "", 
         35000, 7000, "Pedro Martínez", "Feliz cumpleaños", "Ana",
         "Los Militares 2345, Las Condes", "Las Condes", "Cumpleaños", "Archivado", "Pagado", "Normal", "", ""],
        
        # Isabel Torres (CLI007, +56945678901) - Fiel - 6 pedidos
        ["PED008", "2025-10-23 10:00", "2025-10-24 09:00", "JUEVES", "WhatsApp", "", 
         "Isabel Torres", "+56945678901", "Dulce Lirio", "Sin eucalipto por favor", 
         48000, 10000, "Laura Torres", "Para la mejor mamá del mundo", "Isabel",
         "Av. Los Leones 1234, Providencia", "Providencia", "Día de la Madre", "Entregas de Hoy", "No Pagado", "Normal", "", ""],
        
        ["PED009", "2025-09-18 15:30", "2025-09-19 11:00", "MIERCOLES", "Shopify", "#SH1180", 
         "Isabel Torres", "+56945678901", "Lirios Blancos", "Le gustan mucho", 
         32000, 10000, "Roberto Torres", "Aniversario feliz", "Tu esposo",
         "Av. Los Leones 1234, Providencia", "Providencia", "Aniversario", "Archivado", "Pagado", "Normal", "", ""],
        
        # Valentina Rojas (CLI009, +56956789012) - Fiel - 7 pedidos
        ["PED010", "2025-10-23 12:30", "2025-10-25 15:00", "VIERNES", "Shopify", "#SH1237", 
         "Valentina Rojas", "+56956789012", "Amor Eterno", "Incluir tarjeta romántica", 
         65000, 7000, "Sebastián Rojas", "Eres el amor de mi vida", "Valentina",
         "Av. Las Condes 2345, Las Condes", "Las Condes", "San Valentín", "Entregas para Mañana", "Pagado", "EVENTO", "FACTURA 2346 TR. 25/10/25", ""],
        
        # Diego Fernández (CLI010, +56921098765) - Ocasional - 2 pedidos
        ["PED011", "2025-10-23 16:20", "2025-10-24 19:00", "JUEVES", "WhatsApp", "", 
         "Diego Fernández", "+56921098765", "24 rosas rojas en ramo", "URGENTE - Propuesta matrimonio", 
         55000, 10000, "Camila Torres", "¿Quieres casarte conmigo?", "Diego",
         "Av. Providencia 8901, Providencia", "Providencia", "Propuesta", "Pedidos Semana", "Falta Boleta o Factura", "Normal", "", ""],
        
        # Carolina López (CLI005, +56934567890) - Ocasional - 3 pedidos
        ["PED012", "2025-10-23 15:45", "2025-10-24 13:00", "JUEVES", "Shopify", "#SH1238", 
         "Carolina López", "+56934567890", "Elegancia Rosa", "", 
         38000, 15000, "Isabel López", "Gracias por todo", "Carolina",
         "Av. Kennedy 3456, Vitacura", "Vitacura", "Agradecimiento", "Entregas de Hoy", "No Pagado", "Normal", "", ""],
        
        # Roberto Díaz (CLI006, +56909876543) - VIP - 12 pedidos (mostramos 2)
        ["PED013", "2025-10-23 17:00", "2025-10-26 12:00", "SABADO", "Shopify", "#SH1239", 
         "Roberto Díaz", "+56909876543", "Arreglo Corporativo Mensual", "Recepción empresa", 
         80000, 7000, "", "", "",
         "Av. El Bosque 7890, Las Condes", "Las Condes", "Corporativo", "Pedidos Semana", "Pagado", "MANTENCIONES", "", ""],
        
        ["PED014", "2025-10-01 10:00", "2025-10-02 09:00", "MARTES", "Shopify", "#SH1150", 
         "Roberto Díaz", "+56909876543", "Centro de Mesa Ejecutivo", "Sala de reuniones", 
         75000, 7000, "", "", "",
         "Av. El Bosque 7890, Las Condes", "Las Condes", "Corporativo", "Archivado", "Pagado", "MANTENCIONES", "", ""],
        
        # Pedro Silva (CLI004, +56998765432) - Nuevo - 1 pedido
        ["PED015", "2025-10-24 09:00", "2025-10-25 14:00", "VIERNES", "WhatsApp", "", 
         "Pedro Silva", "+56998765432", "Ramo de Bienvenida", "Primer pedido", 
         35000, 15000, "Ana Silva", "Bienvenida", "Pedro",
         "Av. Las Rejas 456, Estación Central", "Estación Central", "Otro", "Pedido", "No Pagado", "Normal", "", ""],
    ]
    
    for fila in pedidos:
        ws.append(fila)
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 18
    ws.column_dimensions['N'].width = 30
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 35
    ws.column_dimensions['Q'].width = 18  # Comuna
    ws.column_dimensions['R'].width = 15  # Motivo
    ws.column_dimensions['S'].width = 18  # Estado
    ws.column_dimensions['T'].width = 15  # Estado Pago
    ws.column_dimensions['U'].width = 15  # Tipo Pedido
    ws.column_dimensions['V'].width = 20  # Cobranza
    ws.column_dimensions['W'].width = 25  # Foto Enviado
    
    wb.save("04_Pedidos.xlsx")
    print("✅ Archivo '04_Pedidos.xlsx' creado")

def crear_recetas_productos():
    """Crea archivo con las 'recetas' de cada producto (qué insumos necesita)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Recetas"
    
    # Encabezados
    headers = ["ID Producto", "Nombre Producto", "ID Insumo", "Tipo Insumo", "Descripción Insumo", "Cantidad Necesaria", "Unidad"]
    ws.append(headers)
    crear_estilo_header(ws)
    
    # Recetas demo
    recetas = [
        # Pasión Roja (PR001)
        ["PR001", "Pasión Roja", "FL001", "Flor", "Rosa Roja", "12", "Tallos"],
        ["PR001", "Pasión Roja", "FL019", "Flor", "Eucalipto Verde", "5", "Ramas"],
        ["PR001", "Pasión Roja", "CO002", "Contenedor", "Florero Vidrio Mediano", "1", "Unidad"],
        
        # Sueño Blanco (PR002)
        ["PR002", "Sueño Blanco", "FL002", "Flor", "Rosa Blanco", "10", "Tallos"],
        ["PR002", "Sueño Blanco", "FL005", "Flor", "Lirio Blanco", "2", "Tallos"],
        ["PR002", "Sueño Blanco", "FL019", "Flor", "Eucalipto Verde", "3", "Ramas"],
        ["PR002", "Sueño Blanco", "CO002", "Contenedor", "Florero Vidrio Mediano", "1", "Unidad"],
        
        # Jardín Primaveral (PR003)
        ["PR003", "Jardín Primaveral", "FL015", "Flor", "Gerbera Naranja", "5", "Tallos"],
        ["PR003", "Jardín Primaveral", "FL016", "Flor", "Gerbera Rosado", "5", "Tallos"],
        ["PR003", "Jardín Primaveral", "FL018", "Flor", "Alstroemeria Morado", "5", "Tallos"],
        ["PR003", "Jardín Primaveral", "FL020", "Flor", "Solidago Amarillo", "3", "Ramas"],
        ["PR003", "Jardín Primaveral", "CO001", "Contenedor", "Florero Vidrio Grande", "1", "Unidad"],
        
        # Elegancia Rosa (PR004)
        ["PR004", "Elegancia Rosa", "FL003", "Flor", "Rosa Rosado", "12", "Tallos"],
        ["PR004", "Elegancia Rosa", "FL019", "Flor", "Eucalipto Verde", "4", "Ramas"],
        ["PR004", "Elegancia Rosa", "CO004", "Contenedor", "Florero Cerámica Blanco", "1", "Unidad"],
        
        # Sol Radiante (PR005)
        ["PR005", "Sol Radiante", "FL007", "Flor", "Girasol Amarillo", "5", "Tallos"],
        ["PR005", "Sol Radiante", "FL015", "Flor", "Gerbera Naranja", "3", "Tallos"],
        ["PR005", "Sol Radiante", "FL020", "Flor", "Solidago Amarillo", "5", "Ramas"],
        ["PR005", "Sol Radiante", "CO002", "Contenedor", "Florero Vidrio Mediano", "1", "Unidad"],
        
        # Dulce Lirio (PR006)
        ["PR006", "Dulce Lirio", "FL005", "Flor", "Lirio Blanco", "4", "Tallos"],
        ["PR006", "Dulce Lirio", "FL006", "Flor", "Lirio Rosado", "3", "Tallos"],
        ["PR006", "Dulce Lirio", "FL019", "Flor", "Eucalipto Verde", "3", "Ramas"],
        ["PR006", "Dulce Lirio", "CO001", "Contenedor", "Florero Vidrio Grande", "1", "Unidad"],
        
        # Orquídea Imperial (PR008)
        ["PR008", "Orquídea Imperial", "FL017", "Flor", "Orquídea Blanco", "3", "Tallos"],
        ["PR008", "Orquídea Imperial", "CO007", "Contenedor", "Macetero Cerámica Mediano", "1", "Unidad"],
    ]
    
    for fila in recetas:
        ws.append(fila)
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 10
    
    wb.save("05_Recetas_Productos.xlsx")
    print("✅ Archivo '05_Recetas_Productos.xlsx' creado")

def crear_proveedores():
    """Crea archivo de proveedores"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    # Encabezados
    headers = ["ID", "Nombre", "Contacto", "Teléfono", "Email", "Especialidad", "Días Entrega", "Notas"]
    ws.append(headers)
    crear_estilo_header(ws)
    
    # Proveedores demo
    proveedores = [
        ["PROV001", "Flores del Valle", "Juan Pérez", "+56987654321", "contacto@floresdelval.cl", "Rosas, Claveles, Gerberas", "Lunes, Miércoles, Viernes", "Entrega temprano"],
        ["PROV002", "Jardín Central", "María Silva", "+56987654322", "ventas@jardincentral.cl", "Lirios, Tulipanes, Orquídeas", "Martes, Jueves", "Flores premium"],
        ["PROV003", "Campo Florido", "Carlos Ramírez", "+56987654323", "info@campoflorido.cl", "Girasoles, Hortensias, Follaje", "Lunes, Jueves", "Productos de temporada"],
        ["PROV004", "Cerámica Artesanal", "Ana Torres", "+56987654324", "ventas@ceramicaarte.cl", "Floreros, Maceteros", "A pedido", "Mínimo 5 unidades"],
        ["PROV005", "Mimbrería Los Andes", "Luis González", "+56987654325", "contacto@mimbreria.cl", "Canastos, Cestas", "A pedido", "Productos artesanales"],
    ]
    
    for fila in proveedores:
        ws.append(fila)
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 20
    
    wb.save("06_Proveedores.xlsx")
    print("✅ Archivo '06_Proveedores.xlsx' creado")

def crear_clientes():
    """Crea archivo con base de datos de clientes"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Encabezados
    headers = ["ID", "Nombre", "Teléfono", "Email", "Tipo Cliente", "Dirección Principal", "Notas", "Total Pedidos", "Total Gastado"]
    ws.append(headers)
    crear_estilo_header(ws)
    
    # Clientes demo
    clientes = [
        ["CLI001", "María González", "+56912345678", "maria.gonzalez@email.com", "Fiel", "Av. Apoquindo 1234, Las Condes", "Cliente frecuente, prefiere rosas rojas", "8", "280000"],
        ["CLI002", "Juan Pérez", "+56987654321", "juan.perez@email.com", "VIP", "Av. Vitacura 5678, Vitacura", "Cliente corporativo, pedidos grandes", "15", "850000"],
        ["CLI003", "Ana Martínez", "+56923456789", "ana.martinez@email.com", "Cumplidor", "Los Militares 2345, Las Condes", "Siempre paga a tiempo", "5", "175000"],
        ["CLI004", "Pedro Silva", "+56998765432", "", "Nuevo", "", "Primer pedido", "1", "35000"],
        ["CLI005", "Carolina López", "+56934567890", "carolina.lopez@email.com", "Ocasional", "Av. Kennedy 3456, Vitacura", "Solo para fechas especiales", "3", "120000"],
        ["CLI006", "Roberto Díaz", "+56909876543", "roberto.diaz@empresa.cl", "VIP", "Av. El Bosque 7890, Las Condes", "Empresa, pedidos mensuales", "12", "620000"],
        ["CLI007", "Isabel Torres", "+56945678901", "isabel.torres@email.com", "Fiel", "Av. Los Leones 1234, Providencia", "Le gustan los lirios", "6", "195000"],
        ["CLI008", "Francisco Morales", "+56910987654", "", "No Cumplidor", "Av. Grecia 5678, Ñuñoa", "Problemas con pagos anteriores", "2", "70000"],
        ["CLI009", "Valentina Rojas", "+56956789012", "valentina.rojas@email.com", "Fiel", "Av. Las Condes 2345, Las Condes", "Cumpleaños el 15 de marzo", "7", "245000"],
        ["CLI010", "Diego Fernández", "+56921098765", "diego.fernandez@email.com", "Ocasional", "Av. Providencia 8901, Providencia", "Aniversario en junio", "2", "85000"],
        ["CLI011", "Camila Soto", "+56967890123", "camila.soto@email.com", "Cumplidor", "Av. Vitacura 3456, Vitacura", "Prefiere arreglos modernos", "4", "156000"],
        ["CLI012", "Andrés Castro", "+56932109876", "", "Nuevo", "", "", "0", "0"],
        ["CLI013", "Sofía Muñoz", "+56978901234", "sofia.munoz@email.com", "Fiel", "Av. Isidora 6789, Las Condes", "Cliente desde hace 2 años", "9", "315000"],
        ["CLI014", "Matías Herrera", "+56943210987", "matias.herrera@empresa.cl", "VIP", "Av. Andrés Bello 4567, Providencia", "Eventos corporativos", "11", "580000"],
        ["CLI015", "Javiera Reyes", "+56989012345", "javiera.reyes@email.com", "Ocasional", "Av. Manquehue 7890, Vitacura", "San Valentín y cumpleaños", "3", "105000"],
    ]
    
    for cliente in clientes:
        ws.append(cliente)
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    wb.save("07_Clientes.xlsx")
    print("✅ Archivo '07_Clientes.xlsx' creado")

if __name__ == "__main__":
    print("\n🌸 Generando archivos Excel demo para Las-Lira...\n")
    
    crear_inventario_flores()
    crear_inventario_contenedores()
    crear_productos_catalogo()
    crear_pedidos()
    crear_recetas_productos()
    crear_proveedores()
    crear_clientes()
    
    print("\n✨ ¡Todos los archivos han sido creados exitosamente!\n")
    print("Archivos generados:")
    print("  1. 01_Inventario_Flores.xlsx")
    print("  2. 02_Inventario_Contenedores.xlsx")
    print("  3. 03_Catalogo_Productos.xlsx")
    print("  4. 04_Pedidos.xlsx")
    print("  5. 05_Recetas_Productos.xlsx")
    print("  6. 06_Proveedores.xlsx")
    print("  7. 07_Clientes.xlsx")
    print("\n📁 Los archivos están listos para subir a Google Drive")

